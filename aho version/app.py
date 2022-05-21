import csvUtil
import openpyxl
import os
from enum import Enum
from pathlib import Path

class LogStatus(Enum):
    read=7
    write=8
    process=9

def logToXLSX(record, i):
    path = Path(os.getcwd())
    print('logging')
    base_row = 2
    wb2 = openpyxl.load_workbook(os.path.join(path.parent.absolute(),"benchmark_sharaf.xlsx"))
    sheet = wb2.active
    print(record)
    print(i)
    if(record['type']=='read'):
        sheet.cell(row=base_row+i, column = LogStatus.read.value).value=record['value']
    elif(record['type']=='write'):
        sheet.cell(row=base_row+i, column = LogStatus.write.value).value=record['value']
    elif(record['type']=='process'):
        sheet.cell(row=base_row+i, column = LogStatus.process.value).value=record['value']
    wb2.save(os.path.join(path.parent.absolute(),"benchmark_sharaf.xlsx"))
for i in range(1,31):
    values = csvUtil.main(i,i*10000,i*1000)
    
    logToXLSX({'type':'read','value':values[0]},i)
    logToXLSX({'type':'write','value':values[1]},i)
    logToXLSX({'type':'process','value':values[2]},i)
